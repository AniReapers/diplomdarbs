import pythoncom
import re
import os
import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from win32com.client import constants
import win32com.client as win32

root = tk.Tk()
root.withdraw()
input_docx = filedialog.askopenfilename(
    title="Izvēlieties DOCX failu",
    filetypes=[("Word dokumenti", "*.docx")]
)
if not input_docx:
    print("Nav izvēlēts fails, iziešana.")
    exit()

output_xlsx = filedialog.asksaveasfilename(
    title="Saglabāt Excel failu kā...",
    defaultextension=".xlsx",
    filetypes=[("Excel faili", "*.xlsx")],
    initialfile="Atskaites_izvade.xlsx"
)
if not output_xlsx:
    print("Nav norādīts faila nosaukums, iziešana.")
    exit()

DOCX_PATH = input_docx
OUTPUT_XLSX = output_xlsx

ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
number_format = r'_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

SECTIONS = {
    "Vadības ziņojums": "Management letter",
    "Peļņas vai zaudējumu": "FACE",
    "Bilance": "FACE",
    "Naudas plūsmas": "CF",
    "Pašu kapitāla": "EMT",
    "Finanšu pārskata pielikums": "Notes",
}

SUMMARY_PATTERNS = [
    re.compile(r"kopā[\s:.\-]*$", re.IGNORECASE),
    re.compile(r"summa[\s:.\-]*$", re.IGNORECASE),
    re.compile(r"kopsumma[\s:.\-]*$", re.IGNORECASE),
    re.compile(r".*31\. decembrī$", re.IGNORECASE),
    re.compile(r".*31\. janvārī$", re.IGNORECASE)
]

NOTE_REGEX = re.compile(r"^([3-9]|[1-2][0-9]|30)\.\s*(.+)?")

SECTION_KEYWORDS = {
    "FACE": [
        "peļņas vai zaudējumu aprēķins",
        "neto apgrozījums",
        "bruto peļņa vai zaudējumi",
        "aktīvs",
        "ilgtermiņa ieguldījumi",
        "nemateriālie ieguldījumi",
        "attīstības izmaksas",
        "patenti",
        "bilance",
        "pasīvs",
        "daļu kapitāls",
        "rezerves",
        "pašu kapitāls"
    ],
    "CF": [
        "pamatdarbības naudas plūsma",
        "bruto pamatdarbības naudas plūsma",
        "naudas plūsma",
        "izdevumi procentu maksājumiem"
    ],
    "EMT": [
        "pašu kapitāla izmaiņu pārskats",
        "kapitāla izmaiņu",
        "emisijas uzcenojums",
        "nesadalītā peļņa",
        "dividenžu izmaksa",
        "iepriekšējā gada peļņas sadale",
        "peļņas sadale",
        "rezervju atlikuma",
        "palielinājums",
        "samazinājums"
    ],
    "IAMT": [
        "nemateriāliem ieguldījumiem",
        "licences",
        "preču zīmes",
        "koncesijas",
        "bilances vērtība",
        "amortizācija"
    ],
    "FAMT": [
        "zeme",
        "ēkas un būves",
        "tehnoloģiskās iekārtas",
        "pamatlīdzekļi",
        "inventārs",
        "bilances vērtība",
        "nolietojums"
    ],
    "Notes": [f"{i}." for i in range(3, 31)]
}

def extract_docx_text_tables(path):
    pythoncom.CoInitialize()
    word = win32.gencache.EnsureDispatch("Word.Application")
    word.Visible = False
    doc = word.Documents.Open(os.path.abspath(path))

    sections = {title: [] for title in SECTIONS.values()}
    sections["IAMT"] = []
    sections["FAMT"] = []

    current = None
    started = False
    management_started = False
    started_notes = False
    skipping_footer = False

    paragraphs = list(doc.Paragraphs)
    tables = list(doc.Tables)
    para_index = 0
    table_index = 0

    def get_next_item():
        next_para = paragraphs[para_index] if para_index < len(paragraphs) else None
        next_table = tables[table_index] if table_index < len(tables) else None

        if not next_para:
            return ('t', next_table)
        if not next_table:
            return ('p', next_para)
        return ('p', next_para) if next_para.Range.Start < next_table.Range.Start else ('t', next_table)

    while para_index < len(paragraphs) or table_index < len(tables):
        kind, item = get_next_item()

        if kind == 'p':
            para = item
            para_index += 1
            if para.Range.Information(constants.wdWithInTable):
                continue

            range_obj = para.Range
            text = range_obj.Text.strip()

            if para.Range.ListFormat.ListType != 0:
                try:
                    number_text = para.Range.ListFormat.ListString
                    text = f"{number_text} {text}"
                except Exception:
                    pass

            if not text:
                continue

            lower_text = text.lower()
            if lower_text.startswith("pielikums ir šī finanšu pārskata"):
                skipping_footer = True
                continue
            elif lower_text.startswith("šis dokuments ir elektroniski parakstīts"):
                skipping_footer = False
                continue
            if skipping_footer:
                continue

            if not started and text == "Vadības ziņojums":
                started = True
                management_started = True
                current = "Management letter"
                continue

            if not started:
                continue

            if text.startswith("Peļņas vai zaudējumu"):
                current = "FACE"
                management_started = False
            elif text.startswith("Bilance"):
                current = "FACE"
                management_started = False
            elif text.startswith("Naudas plūsmas"):
                current = "CF"
                management_started = False
            elif text.startswith("Pašu kapitāla"):
                current = "EMT"
                management_started = False
            elif text.startswith("Finanšu pārskata pielikums"):
                current = "Notes"
                started_notes = True
                management_started = False
            elif started_notes:
                current = "Notes"
            elif management_started:
                current = "Management letter"

            if current:
                sections[current].append(text)

        elif kind == 't':
            table = item
            table_index += 1

            if not started or not current or skipping_footer:
                continue

            for row in table.Rows:
                row_data = []
                for cell in row.Cells:
                    cell_text = cell.Range.Text.strip().replace('\r\x07', '')
                    row_data.append(cell_text)
                sections[current].append('\t'.join(row_data))

    doc.Close(False)
    word.Quit()
    pythoncom.CoUninitialize()
    return sections

def apply_styles(cell, wrap_text=False):
    cell.font = Font(name="Arial Narrow", size=10)
    alignment = Alignment(wrap_text=wrap_text, vertical="top", horizontal="left")
    cell.alignment = alignment

def clean_text(text):
    return ILLEGAL_CHARACTERS_RE.sub("", text)
    
def insert_text(ws, lines, start_row):
    wrap_text = False
    if ws.title == "Management letter":
        ws.column_dimensions["A"].width = 80
        wrap_text = True
    elif ws.title == "Notes":
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 50
        for col_letter in ["C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = 17
        wrap_text = True

    row_pointer = start_row
    for line in lines:
        line = clean_text(line)

        is_kopa_row = False

        if '\t' in line:
            values = line.split('\t')
            is_kopa_row = any("kopā" in str(val).lower() for val in values[0:5])

            for col_idx, cell in enumerate(values, start=1):
                c = ws.cell(row=row_pointer, column=col_idx, value=cell)
                apply_styles(c, wrap_text)
                if is_kopa_row:
                    c.font = Font(name="Arial Narrow", size=10, bold=True)
            if is_kopa_row:
                row_pointer += 1
                ws.insert_rows(row_pointer, amount=3)
                recalc_cell = ws.cell(row=row_pointer, column=1, value="Recalculated")
                diff_cell = ws.cell(row=row_pointer + 1, column=1, value="Difference")
                apply_styles(recalc_cell, wrap_text)
                apply_styles(diff_cell, wrap_text)
                recalc_cell.font = Font(name="Arial Narrow", size=10, color="0070C0", bold=True)
                diff_cell.font = Font(name="Arial Narrow", size=10, color="FF0000", bold=True)
                row_pointer += 3
            else:
                row_pointer += 1
        else:
            c = ws.cell(row=row_pointer, column=1, value=line)
            apply_styles(c, wrap_text)
            row_pointer += 1

    return row_pointer + 1

def is_year_cell(value):
    try:
        year = int(str(value).strip())
        return 1900 <= year <= 2100
    except:
        return False

def is_string_cell(value):
    return isinstance(value, str) and value.strip() != "" and not value.strip().isdigit()

def generate_recalculated_and_difference_formulas(ws):
    found_any = False
    for row in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row, column=2)
        if label_cell.value and str(label_cell.value).strip().lower() == "recalculated":
            print(f"Found 'Recalculated' at row {row}")
            found_any = True
            for col in range(3, 5):  # C and D only
                col_letter = get_column_letter(col)
                target_cell = ws.cell(row=row, column=col)

                header_row = row - 1
                while header_row > 1:
                    val = ws.cell(row=header_row, column=col).value
                    if val is None or (isinstance(val, str) and val.strip() != "" and not any(char.isdigit() for char in val)) or is_year_cell(val):
                        break
                    header_row -= 1

                total_row = row - 1
                while total_row > header_row:
                    val = ws.cell(row=total_row, column=2).value
                    if isinstance(val, str) and "kopā" in val.lower():
                        break
                    total_row -= 1

                if is_year_cell(ws.cell(row=header_row, column=col).value):
                    sum_start = header_row + 1
                else:
                    sum_start = header_row + 1
                sum_end = total_row - 1

                if sum_start <= sum_end:
                    sum_range = f"{col_letter}{sum_start}:{col_letter}{sum_end}"
                    print(f"Setting formula in {col_letter}{row}: =SUM({sum_range})")
                    target_cell.value = f"=SUM({sum_range})"
                    target_cell.font = Font(name="Arial Narrow", size=10, color="0070C0", bold=True)
                    diff_cell = ws.cell(row=row + 1, column=col)
                    diff_label = ws.cell(row=row + 1, column=2)
                    if diff_label.value and str(diff_label.value).strip().lower() == "difference":
                        orig_total = ws.cell(row=total_row, column=col)
                        diff_cell.value = f"={orig_total.coordinate}-{target_cell.coordinate}"
                        diff_cell.font = Font(name="Arial Narrow", size=10, color="FF0000", bold=True)
                else:
                    print(f"Skipped formula in {col_letter}{row}: range {sum_start}-{sum_end} invalid")

    if not found_any:
        print("No 'Recalculated' labels found in column B.")

def debug_recalculated_neighbors(ws):
    for row in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row, column=2)
        if label_cell.value and str(label_cell.value).strip().lower() == "recalculated":
            print(f"Row {row} Recalculated:")
            for col in range(3, 5):
                print(f"  {get_column_letter(col)}{row-1}: {ws.cell(row=row-1, column=col).value}")
                print(f"  {get_column_letter(col)}{row-2}: {ws.cell(row=row-2, column=col).value}")
                print(f"  {get_column_letter(col)}{row-3}: {ws.cell(row=row-3, column=col).value}")
                print(f"  {get_column_letter(col)}{row-4}: {ws.cell(row=row-4, column=col).value}")

def is_summary_row(cell_val):
    if not isinstance(cell_val, str):
        return False
    return any(pat.match(cell_val.strip()) for pat in SUMMARY_PATTERNS)

def fix_latvian_numbers(cell_value):
    if isinstance(cell_value, str):
        s = cell_value.strip().replace(" ", "").replace("\xa0", "")
        if s == "-":
            return 0
        if re.match(r'^-?\(?\d{1,3}(\d{3})*(,\d+)?\)?$', s):
            s = s.replace("(", "-").replace(")", "").replace(",", ".")
            try:
                return float(s)
            except ValueError:
                return cell_value
    return cell_value

def fix_number_strings_in_workbook(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                original_value = cell.value
                new_value = fix_latvian_numbers(original_value)
                if isinstance(new_value, (int, float)):
                    cell.value = new_value
                    cell.number_format = number_format
                elif isinstance(original_value, str) and original_value.startswith("="):
                    cell.number_format = number_format
                elif isinstance(original_value, (int, float)):
                    cell.number_format = number_format

if __name__ == "__main__":
    print("Extracting DOCX text & Word tables...")
    text_sections = extract_docx_text_tables(DOCX_PATH)

    print("Building output Excel workbook...")
    workbook = Workbook()
    workbook.remove(workbook.active)
    for section, lines in text_sections.items():
        ws = workbook.create_sheet(title=section[:31])
        insert_text(ws, lines, 1)
        if section == "Notes":
            ws.insert_cols(1)
            ws.column_dimensions["A"].width = 2
            generate_recalculated_and_difference_formulas(ws)
            debug_recalculated_neighbors(ws)

    print("Post-processing numbers...")
    fix_number_strings_in_workbook(workbook)

    print("Saving workbook...")
    workbook.save(OUTPUT_XLSX)
    print(f"Done! Saved as: {OUTPUT_XLSX}")

